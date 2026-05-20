# views.py
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User as DjangoUser
from django.db import IntegrityError
from django.http import HttpResponse
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect, get_object_or_404
from django.db import transaction 

import io
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from .forms import UserEditForm

# Modelos y serializadores
# Modelos y serializadores
from user.models import Users
from user.serializers import UserSerializer
from rest_framework.generics import ListAPIView
from django_filters.rest_framework import DjangoFilterBackend


# Telegram service (usando requests)
from telegram_service import send_telegram_message, send_telegram_document
# notificaciones 
from django.contrib import messages
from django.utils.timezone import make_aware
from datetime import datetime



# --- API para listar usuarios ---
class UserListView(ListAPIView):
    queryset = Users.objects.all()
    serializer_class = UserSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ["username", "id", "email"] 


# --- Autenticación ---
def login_view(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user:
            login(request, user)
            return redirect('home')
        else:
            context['error'] = 'Credenciales inválidas. Intenta de nuevo.'

    return render(request, 'login.html', context)

#--- HOME -----
@login_required
def home(request):
    """Home con rol del usuario"""
    from django.utils.timezone import localdate
    from clases.models import Clase, Inscripcion

    # Buscar el perfil extendido por username
    try:
        perfil = Users.objects.get(username=request.user.username)
        role_name = perfil.role.nombre if perfil.role else None
    except Users.DoesNotExist:
        perfil = None
        role_name = None  # Si por alguna razón no tiene perfil

    # Estadísticas dinámicas según el rol
    usuarios_totales = 0
    clases_hoy = 0
    ingresos_totales = 0
    clases_inscritas_totales = 0
    proximas_clases_24h = 0
    clases_entrenadas_totales = 0
    alumnos_totales = 0

    hoy = localdate()

    if role_name == "Administrador":
        usuarios_totales = Users.objects.count()
        clases_hoy = Clase.objects.filter(fecha=hoy).count()
        clientes_activos = Users.objects.filter(role__nombre="Cliente", estado="Activo").count()
        ingresos_totales = clientes_activos * 80000  # $80,000 COP por cada cliente activo

    elif role_name == "Cliente":
        clases_inscritas_totales = Inscripcion.objects.filter(cliente=request.user).count()
        # Próximas clases hoy en adelante
        proximas_clases_24h = Inscripcion.objects.filter(cliente=request.user, clase__fecha__gte=hoy).count()

    elif role_name == "Entrenador":
        clases_entrenadas_totales = Clase.objects.filter(entrenador=request.user).count()
        # Sumar los alumnos inscritos en todas las clases de este entrenador
        alumnos_totales = Inscripcion.objects.filter(clase__entrenador=request.user).count()

    return render(request, "home.html", {
        "role_name": role_name,
        "user": request.user,
        "perfil": perfil,
        "usuarios_totales": usuarios_totales,
        "clases_hoy": clases_hoy,
        "ingresos_totales": ingresos_totales,
        "clases_inscritas_totales": clases_inscritas_totales,
        "proximas_clases_24h": proximas_clases_24h,
        "clases_entrenadas_totales": clases_entrenadas_totales,
        "alumnos_totales": alumnos_totales,
    })

def signup_view(request):
    if request.method == 'POST':
        try:
            username = request.POST['username']
            name = request.POST['name']
            last_name = request.POST['last_name']
            email = request.POST['email']
            chat_id = request.POST['chat_id']
            role = request.POST['role']
            password1 = request.POST['password1']
            password2 = request.POST['password2']

            if password1 != password2:
                return render(request, 'signup.html', {'error': 'Las contraseñas no coinciden.'})

            # 1. Validar en la tabla nativa de Django
            if DjangoUser.objects.filter(username=username).exists():
                return render(request, 'signup.html', {'error': 'El nombre de usuario ya está en uso en el sistema.'})

            if DjangoUser.objects.filter(email=email).exists():
                return render(request, 'signup.html', {'error': 'El correo electrónico ya está registrado.'})

            # 2. VALIDACIÓN CLAVE: Validar también en tu tabla personalizada 'Users' antes de guardar
            if Users.objects.filter(username=username).exists():
                return render(request, 'signup.html', {'error': 'El nombre de usuario ya existe en los perfiles de GymPower.'})
            
            if chat_id and Users.objects.filter(chat_id=chat_id).exists():
                return render(request, 'signup.html', {'error': 'Este Chat ID de Telegram ya está asignado a otro usuario.'})

            # 3. Guardar de forma atómica (O se guardan ambos bien, o no se guarda ninguno)
            with transaction.atomic():
                # Crear usuario de Django
                user = DjangoUser.objects.create_user(username=username, email=email, password=password1)
                user.first_name = name
                user.last_name = last_name
                user.save()

                # Crear perfil extendido en GymPower
                perfil = Users.objects.create(
                    username=username,
                    password=password1,  
                    email=email,
                    chat_id=chat_id,
                    first_name=name,
                    last_name=last_name,
                    role_id=role
                )

            # Notificación Telegram (fuera de la transacción por si falla la red)
            if chat_id:
                try:
                    send_telegram_message(
                        chat_id,
                        f"👋 ¡Hola {name}!\n\n"
                        "Bienvenido a *GymPower* \n"
                        "Tu cuenta se ha creado con éxito.\n\n"
                        "Ya puedes iniciar sesión en la plataforma."
                    )
                except Exception:
                    pass # Que no se rompa el registro si Telegram falla temporalmente

            # Si todo salió perfecto, iniciamos sesión y mandamos al HOME, no al login
            login(request, user)
            return redirect('home')

        except IntegrityError as e:
            return render(request, 'signup.html', {'error': f'Error de duplicidad en la base de datos: {str(e)}'})
        except Exception as e:
            return render(request, 'signup.html', {'error': f'Error durante el registro: {str(e)}'})

    return render(request, 'signup.html')




@login_required
def editar_perfil(request, user_id):
    usuario = get_object_or_404(Users, id=user_id)

    try:
        perfil = Users.objects.get(username=request.user.username)
        role_name = perfil.role.nombre if perfil.role else None
    except Users.DoesNotExist:
        perfil = None
        role_name = None

    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=usuario)
        if form.is_valid():
            form.save()
            return redirect('home')  
    else:
        form = UserEditForm(instance=usuario)

    return render(request, 'editar_usuario.html', {
        'form': form,
        'usuario': usuario,
        'perfil': perfil,
        'role_name': role_name
    })

@login_required
def listar_usuarios(request):
    user = request.user 
    try:
        perfil = Users.objects.get(username=user.username)
        role_name = perfil.role.nombre if perfil.role else None
    except Users.DoesNotExist:
        perfil = None
        role_name = None

    usuarios = Users.objects.all()

    return render(request, 'listar_usuarios.html', {
        'usuarios': usuarios,
        'role_name': role_name,
        'perfil': perfil,
    })

@login_required
def eliminar_usuario(request, user_id):
    current_user = Users.objects.get(username=request.user.username)
    if current_user.role.nombre != "Administrador":
        messages.error(request, "No tienes permiso para eliminar usuarios.")
        return redirect('listar_usuarios')

    usuario = get_object_or_404(Users, id=user_id)
    usuario.delete()
    messages.success(request, f"El usuario {usuario.username} fue eliminado correctamente.")
    return redirect('listar_usuarios')


# --- Reportes PDF ---
@login_required
def generar_pdf(request):
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer)

    p.setFont("Helvetica-Bold", 16)
    p.drawString(200, 800, "📋 Reporte de Usuarios del Gym")

    p.setFont("Helvetica", 10)
    y = 750
    for user in Users.objects.all():
        texto = f"{user.first_name} {user.last_name} | {user.email} | Rol: {user.role} | Estado: {user.estado} | Registrado: {user.fecha_registro.strftime('%Y-%m-%d')}"
        p.drawString(50, y, texto)
        y -= 20
        if y < 50:  
            p.showPage()
            y = 800

    p.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.pdf"'
    return response


# --- Reportes Excel ---
@login_required
def generar_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    ws.append(["Nombre", "Apellido", "Email", "Rol", "Estado", "Fecha Registro"])

    for user in Users.objects.all():
        ws.append([
            user.first_name,
            user.last_name,
            user.email,
            str(user.role) if user.role else "N/A",
            user.estado,
            user.fecha_registro.strftime("%Y-%m-%d") if user.fecha_registro else "N/A"
        ])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="reporte_usuarios.xlsx"'

    wb.save(response)
    return response


# --- Enviar reporte por Telegram ---
@login_required
@csrf_exempt
def enviar_reporte_telegram(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        file_type = request.POST.get("file_type", "pdf")

        try:
            user = Users.objects.get(id=user_id)
        except Users.DoesNotExist:
            return HttpResponse("❌ Usuario no encontrado", status=404)

        if not user.chat_id:
            return HttpResponse(f"⚠️ Usuario {user.username} no tiene chat_id", status=400)

        try:
            chat_id_int = int(user.chat_id)
        except ValueError:
            return HttpResponse(f"❌ chat_id inválido para {user.username}", status=400)

        # Generar archivo en memoria
        buffer = io.BytesIO()
        if file_type == "pdf":
            p = canvas.Canvas(buffer)
            p.setFont("Helvetica-Bold", 16)
            p.drawString(200, 800, "📋 Reporte de Usuarios del Gym")
            p.setFont("Helvetica", 10)
            y = 750
            for u in Users.objects.all():
                texto = f"{u.first_name} {u.last_name} | {u.email} | Rol: {u.role} | Estado: {u.estado} | Registrado: {u.fecha_registro.strftime('%Y-%m-%d')}"
                p.drawString(50, y, texto)
                y -= 20
                if y < 50:
                    p.showPage()
                    y = 800
            p.save()
            buffer.seek(0)
            file_name = f"reporte_usuarios_{now().strftime('%Y%m%d')}.pdf"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Usuarios"
            ws.append(["Nombre", "Apellido", "Email", "Rol", "Estado", "Fecha Registro"])
            for u in Users.objects.all():
                ws.append([
                    u.first_name,
                    u.last_name,
                    u.email,
                    str(u.role) if u.role else "N/A",
                    u.estado,
                    u.fecha_registro.strftime("%Y-%m-%d") if u.fecha_registro else "N/A"
                ])
            wb.save(buffer)
            buffer.seek(0)
            file_name = f"reporte_usuarios_{now().strftime('%Y%m%d')}.xlsx"

        # Enviar mensaje
        mensaje = (
            f"👋 Hola {user.first_name}, aquí tienes tu estado:\n\n"
            f"📧 Email: {user.email}\n"
            f"📌 Rol: {user.role}\n"
            f"✅ Estado: {user.estado}\n"
            f"🗓 Registrado: {user.fecha_registro.strftime('%Y-%m-%d') if user.fecha_registro else 'N/A'}\n\n"
            f"📎 Adjunto encontrarás el reporte en formato {file_type.upper()}."
        )

        send_telegram_message(chat_id_int, mensaje)
        send_telegram_document(chat_id_int, buffer, file_name)

        messages.success(request, "✅ Reporte enviado con éxito")
        return redirect("reportes")


@login_required
def reportes_view(request):
    # Obtener el perfil y su rol
    try:
        perfil = Users.objects.get(username=request.user.username)
        role_name = perfil.role.nombre if perfil.role else None
    except Users.DoesNotExist:
        perfil = None
        role_name = None
    
    # Si el usuario no es administrador, lo redirigimos al home
    if not perfil or role_name != "Administrador":
        return redirect('home')  

    # Si es administrador, mostramos los reportes
    users = Users.objects.all()
    return render(request, "reportes.html", {
        "users": users,
        "perfil": perfil,
        "role_name": role_name
    })


@login_required
def enviar_notificacion(request):
    if request.method == "POST":
        destinatario_username = request.POST.get("destinatario")
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        fecha_envio = request.POST.get("fecha_envio")

        try:
            user = Users.objects.get(username=destinatario_username)
        except Users.DoesNotExist:
            messages.error(request, "❌ Usuario no encontrado")
            return redirect("enviar_notificacion")

        if not user.chat_id:
            messages.error(request, f"⚠️ Usuario {user.username} no tiene chat_id")
            return redirect("enviar_notificacion")

        try:
            chat_id_int = int(user.chat_id)
        except ValueError:
            messages.error(request, f"❌ chat_id inválido para {user.username}")
            return redirect("enviar_notificacion")

        # Crear mensaje formateado
        mensaje = (
            f"📢 <b>{titulo}</b>\n\n"
            f"{descripcion}\n\n"
            f"🗓 Fecha de envío: {fecha_envio}"
        )

        send_telegram_message(chat_id_int, mensaje)
        messages.success(request, "✅ Notificación enviada con éxito")
        return redirect("enviar_notificacion")

    usuarios = Users.objects.all()
    return render(request, "enviar_notificacion.html", {"usuarios": usuarios})

from .models import Notificacion

@login_required
def notificaciones_view(request, notificacion_id=None):
    """
    Muestra el formulario y la lista de notificaciones.
    Si hay notificacion_id, carga esa notificación para edición.
    """
    try:
        perfil = Users.objects.get(username=request.user.username)
        role_name = perfil.role.nombre if perfil.role else None
    except Users.DoesNotExist:
        perfil = None
        role_name = None

    notificacion = None
    if notificacion_id:
        notificacion = Notificacion.objects.get(id=notificacion_id)

    if request.method == "POST":
        destinatario_username = request.POST.get("destinatario")
        titulo = request.POST.get("titulo")
        descripcion = request.POST.get("descripcion")
        fecha_envio = request.POST.get("fecha_envio")

        try:
            user = Users.objects.get(username=destinatario_username)
        except Users.DoesNotExist:
            messages.error(request, "❌ Usuario no encontrado")
            return redirect("notificaciones")

        fecha_envio_aware = None
        if fecha_envio:
            try:
                # Handle YYYY-MM-DDTHH:MM or YYYY-MM-DD HH:MM:SS or YYYY-MM-DDTHH:MM:SS
                fecha_envio_str = fecha_envio.replace('T', ' ')
                if len(fecha_envio_str) == 16:
                    dt = datetime.strptime(fecha_envio_str, '%Y-%m-%d %H:%M')
                else:
                    dt = datetime.strptime(fecha_envio_str, '%Y-%m-%d %H:%M:%S')
                fecha_envio_aware = make_aware(dt)
            except ValueError:
                fecha_envio_aware = fecha_envio
        else:
            fecha_envio_aware = None

        if notificacion:  # Editar existente
            notificacion.destinatario = user
            notificacion.titulo = titulo
            notificacion.descripcion = descripcion
            notificacion.fecha_envio = fecha_envio_aware
            notificacion.save()
            messages.success(request, "✏️ Notificación actualizada correctamente")
        else:  # Crear nueva
            Notificacion.objects.create(
                destinatario=user,
                titulo=titulo,
                descripcion=descripcion,
                fecha_envio=fecha_envio_aware
            )
            messages.success(request, "✅ Notificación creada correctamente")

        return redirect("notificaciones")

    notificaciones = Notificacion.objects.select_related("destinatario").order_by("-fecha_creacion")
    usuarios = Users.objects.all()
    return render(request, "enviar_notificaciones.html", {
        "usuarios": usuarios,
        "notificaciones": notificaciones,
        "notificacion": notificacion,
        "perfil": perfil,
        "role_name": role_name
    })


@login_required
def notificacion_delete(request, id):
    notif = Notificacion.objects.get(id=id)
    notif.delete()
    messages.success(request, "🗑️ Notificación eliminada correctamente")
    return redirect("notificaciones")


@login_required
def notificacion_enviar(request, id):
    notif = Notificacion.objects.get(id=id)
    user = notif.destinatario

    if not user.chat_id:
        messages.error(request, f"⚠️ El usuario {user.username} no tiene chat_id configurado.")
        return redirect("notificaciones")

    try:
        chat_id_int = int(user.chat_id)
    except ValueError:
        messages.error(request, f"❌ chat_id inválido para {user.username}")
        return redirect("notificaciones")

    mensaje = (
        f"📢 <b>{notif.titulo}</b>\n\n"
        f"{notif.descripcion}\n\n"
        f"🗓 Fecha programada: {notif.fecha_envio.strftime('%Y-%m-%d %H:%M')}"
    )

    send_telegram_message(chat_id_int, mensaje)
    notif.enviado = True
    notif.save()
    messages.success(request, "✅ Notificación enviada correctamente")
    return redirect("notificaciones")


from django.core.mail import EmailMessage

@login_required
@csrf_exempt
def enviar_reporte_email(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        file_type = request.POST.get("file_type", "pdf")

        try:
            user = Users.objects.get(id=user_id)
        except Users.DoesNotExist:
            return HttpResponse("❌ Usuario no encontrado", status=404)

        if not user.email:
            return HttpResponse(f"⚠️ Usuario {user.username} no tiene correo registrado", status=400)

        buffer = io.BytesIO()
        file_name = f"reporte_usuarios_{now().strftime('%Y%m%d')}"

        # Generar archivo
        if file_type == "pdf":
            p = canvas.Canvas(buffer)
            p.setFont("Helvetica-Bold", 16)
            p.drawString(200, 800, "📋 Reporte de Usuarios del Gym")
            p.setFont("Helvetica", 10)
            y = 750
            for u in Users.objects.all():
                texto = f"{u.first_name} {u.last_name} | {u.email} | Rol: {u.role} | Estado: {u.estado}"
                p.drawString(50, y, texto)
                y -= 20
                if y < 50:
                    p.showPage()
                    y = 800
            p.save()
            buffer.seek(0)
            file_name += ".pdf"
            content_type = "application/pdf"
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Usuarios"
            ws.append(["Nombre", "Apellido", "Email", "Rol", "Estado", "Fecha Registro"])
            for u in Users.objects.all():
                ws.append([u.first_name, u.last_name, u.email, str(u.role) if u.role else "N/A", u.estado, u.fecha_registro.strftime("%Y-%m-%d") if u.fecha_registro else "N/A"])
            wb.save(buffer)
            buffer.seek(0)
            file_name += ".xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

        # Crear email
        subject = "📊 Reporte de Usuarios - GymPower"
        body = f"""
Hola {user.first_name},

Adjunto encontrarás el reporte de usuarios solicitado en formato {file_type.upper()}.

Saludos,
El equipo de GymPower 💪
"""
        email = EmailMessage(
            subject,
            body,
            to=[user.email]
        )
        email.attach(file_name, buffer.getvalue(), content_type)
        email.send()

        messages.success(request, f"✅ Reporte enviado a {user.email}")
        return redirect("reportes")
